"""Tests de humo para el dominio `mining_cochilco`, corridos contra los datos
reales ya descargados/procesados (sin red, sin mocks del pipeline real) --
correr `fetch.py` -> `clean.py` -> `features.py` -> `model.py` antes de estos
tests si los archivos de `data/`/`outputs/` todavía no existen.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from src.domains.mining_cochilco.clean import (
    GRAND_TOTAL_COLUMN,
    RAW_FILENAME,
    SHEET_NAME,
    SUBTOTAL_COLUMNS,
    build_clean_panel,
)
from src.domains.mining_cochilco.features import FEATURE_COLUMNS, SEASONAL_NAIVE_COLUMN, TARGET_COLUMN, build_features

ROOT = Path(__file__).resolve().parents[2]
RAW_PATH = ROOT / "data" / "raw" / "mining" / RAW_FILENAME
PANEL_PATH = ROOT / "data" / "processed" / "mining" / "mining_panel_clean.csv"
FEATURES_PATH = ROOT / "data" / "processed" / "mining" / "mining_features.csv"
METRICS_PATH = ROOT / "outputs" / "mining" / "metrics.json"

pytestmark = pytest.mark.skipif(not RAW_PATH.exists(), reason="run src/domains/mining_cochilco/fetch.py first")

COMPANY_COLUMNS_MIN_COUNT = 35  # COCHILCO ha ido agregando faenas nuevas con el tiempo; >=35 es un piso conservador


@pytest.fixture(scope="module")
def clean_panel() -> pd.DataFrame:
    panel, _report = build_clean_panel()
    return panel


@pytest.fixture(scope="module")
def features_df(clean_panel: pd.DataFrame) -> pd.DataFrame:
    return build_features(clean_panel)


# ---------------------------------------------------------------------------
# Archivo crudo: hoja y estructura esperada
# ---------------------------------------------------------------------------


def test_raw_file_has_expected_sheet_and_header():
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    assert SHEET_NAME in wb.sheetnames

    ws = wb[SHEET_NAME]
    header_row = [ws.cell(row=7, column=c).value for c in range(1, ws.max_column + 1)]
    header_text = " ".join(str(v) for v in header_row if v is not None).lower()
    for token in ["escondida", "collahuasi", "total codelco", "total chile"]:
        assert token in header_text


def test_raw_file_has_nontrivial_row_count():
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    # >150 filas: 6 de titulo + 1 encabezado + >=150 meses reales + resumenes
    # anuales intercalados + fila de cita.
    assert ws.max_row > 150


# ---------------------------------------------------------------------------
# Panel limpio: sin nulos, columnas subtotal validadas, rango plausible
# ---------------------------------------------------------------------------


def test_cleaned_panel_has_no_nulls_in_key_columns(clean_panel: pd.DataFrame):
    key_columns = ["fecha", "total_nacional_miles_ton", "total_codelco_miles_ton", "share_codelco"]
    assert clean_panel[key_columns].isna().sum().sum() == 0


def test_cleaned_panel_company_columns_have_no_nulls(clean_panel: pd.DataFrame):
    # Missingness genuina de este dominio (faena que aun no operaba / ya
    # cerro) se representa como 0.0 explicito en la fuente, no como celda en
    # blanco -- ver clean.py. El panel limpio no deberia tener NINGUN nulo.
    company_cols = [c for c in clean_panel.columns if c not in SUBTOTAL_COLUMNS + [
        GRAND_TOTAL_COLUMN, "fecha", "total_nacional_miles_ton", "total_codelco_miles_ton", "share_codelco",
    ]]
    assert len(company_cols) >= COMPANY_COLUMNS_MIN_COUNT
    assert clean_panel[company_cols].isna().sum().sum() == 0


def test_cleaned_panel_drops_annual_summary_and_footer_rows(clean_panel: pd.DataFrame):
    # Ninguna fila del panel limpio deberia tener una fecha que no sea el
    # primer dia de un mes real (los resumenes anuales usaban un año "pelado"
    # como texto, no una fecha) -- confirma que el filtro de tipo funciono.
    fechas = pd.to_datetime(clean_panel["fecha"])
    assert (fechas.dt.day == 1).all()
    assert fechas.is_monotonic_increasing
    assert fechas.dt.year.min() == 2014


def test_disguised_subtotal_columns_are_excluded_from_national_total(clean_panel: pd.DataFrame):
    # "Chuqui y R.Tomic", "Total Codelco", "Angloamerican Sur" y
    # "Capstone Copper" son subtotales de otras columnas del mismo archivo
    # (confirmado por identidad numerica exacta en clean.py) -- sumar TODAS
    # las columnas de empresa incluyendolas duplicaria produccion real.
    company_cols = [c for c in clean_panel.columns if c not in SUBTOTAL_COLUMNS + [
        GRAND_TOTAL_COLUMN, "fecha", "total_nacional_miles_ton", "total_codelco_miles_ton", "share_codelco",
    ]]
    naive_sum_with_subtotals = clean_panel[company_cols + SUBTOTAL_COLUMNS].sum(axis=1)
    correct_total = clean_panel["total_nacional_miles_ton"]
    # La suma ingenua (con subtotales incluidos) SIEMPRE sobreestima el total
    # real -- si esto fallara, alguna de las 4 columnas dejo de ser subtotal.
    assert (naive_sum_with_subtotals > correct_total).all()


def test_national_total_matches_cochilco_published_total(clean_panel: pd.DataFrame):
    # El total nacional derivado (suma de las 38 columnas reales) debe
    # coincidir, fila a fila, con TOTAL CHILE publicado por COCHILCO --
    # la misma validacion que clean.py corre antes de confiar en la serie.
    deviation = (clean_panel["total_nacional_miles_ton"] - clean_panel[GRAND_TOTAL_COLUMN]).abs()
    assert deviation.max() < 0.01


def test_national_total_within_plausible_real_world_range(clean_panel: pd.DataFrame):
    # Chile produce del orden de 5.0-5.8 millones de toneladas de cobre al
    # año en este periodo -- en miles de T.M. mensuales, eso ubica un mes
    # tipico en un rango amplio pero acotado. Un bug de unidades/columnas
    # daria un numero fuera de orden de magnitud (ej. ~4.6 o ~46000), no uno
    # dentro de este rango.
    total = clean_panel["total_nacional_miles_ton"]
    assert total.between(300, 650).all()
    assert 400 < total.mean() < 550


# ---------------------------------------------------------------------------
# Features: sin look-ahead, forma esperada
# ---------------------------------------------------------------------------


def test_features_have_no_nulls_in_model_columns(features_df: pd.DataFrame):
    required = FEATURE_COLUMNS + [TARGET_COLUMN, SEASONAL_NAIVE_COLUMN]
    assert features_df[required].isna().sum().sum() == 0


def test_features_target_is_shifted_forward_one_month(clean_panel: pd.DataFrame, features_df: pd.DataFrame):
    # target_next_total de la fila de un mes t debe ser exactamente la
    # produccion nacional observada en el mes t+1 -- sin look-ahead ni
    # desalineacion de indice.
    panel_by_date = clean_panel.set_index(pd.to_datetime(clean_panel["fecha"]))["total_nacional_miles_ton"]
    sample = features_df.sample(n=min(10, len(features_df)), random_state=42)
    for _, row in sample.iterrows():
        current_month = pd.Timestamp(row["fecha"])
        next_month = current_month + pd.DateOffset(months=1)
        assert row[TARGET_COLUMN] == pytest.approx(panel_by_date.loc[next_month])


def test_features_row_count_reflects_lag_and_target_loss(clean_panel: pd.DataFrame, features_df: pd.DataFrame):
    # Se pierden 12 filas por el lag mas largo (lag12) y 1 por el target
    # shifted hacia adelante -- exactamente len(panel) - 13 filas utilizables.
    assert len(features_df) == len(clean_panel) - 13


# ---------------------------------------------------------------------------
# Modelo: el mejor modelo real le gana al baseline honesto por un margen real
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not METRICS_PATH.exists(), reason="run src/domains/mining_cochilco/model.py first")
def test_mlp_trained_at_least_100_epochs():
    metrics = json.loads(METRICS_PATH.read_text(encoding="utf-8"))
    assert metrics["results"]["mlp_pytorch"]["epochs_run"] >= 100


@pytest.mark.skipif(not METRICS_PATH.exists(), reason="run src/domains/mining_cochilco/model.py first")
def test_best_model_beats_seasonal_baseline_by_a_real_margin():
    # La afirmacion real de este dominio: con datos oficiales de COCHILCO y
    # un baseline estacional razonablemente dificil de superar, XGBoost
    # captura señal real por sobre "el mismo mes, año anterior" -- se
    # verifica el R2 real obtenido, no se fuerza un resultado.
    metrics = json.loads(METRICS_PATH.read_text(encoding="utf-8"))
    baseline_r2 = metrics["results"]["baseline_estacional"]["r2"]
    xgb_r2 = metrics["results"]["xgboost"]["r2"]
    assert xgb_r2 > baseline_r2 + 0.1
