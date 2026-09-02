"""Descarga el World Development Indicators (WDI) completo del Banco Mundial:
un Excel real de ~80MB, 6 hojas, 401.394 filas indicador×país en la hoja
`Data` -- el tipo de archivo que una consultora recibe de un cliente o de un
organismo público y tiene que transformar en un data warehouse consultable,
no un CSV ya tabular.

No se puede cargar la hoja `Data` completa a un DataFrame de pandas en cada
corrida de desarrollo (leerla entera con openpyxl toma ~50 segundos solo para
iterar, sin contar el parseo) -- por eso este módulo hace un FILTRADO EN
STREAMING fila por fila (openpyxl `read_only=True`, nunca carga todo en
memoria) y solo materializa un subconjunto curado de indicadores reales antes
de escribir a disco. Esto es deliberado y es la técnica central del dominio:
la zona "raw" del data lake conserva el Excel completo tal cual llega del
Banco Mundial; la zona "staging" ya es angosta y manejable.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parents[3]
RAW_DIR = ROOT / "data" / "raw" / "consulting"
ZIP_URL = "https://databankfiles.worldbank.org/public/ddpext_download/WDI_excel.zip"
ZIP_PATH = RAW_DIR / "WDI_excel.zip"
XLSX_PATH = RAW_DIR / "WDIEXCEL.xlsx"

# Indicadores reales curados: predictores plausibles de esperanza de vida al
# nacer (el target del modelo de este dominio) -- elegidos por relevancia de
# dominio (salud, agua/saneamiento, ingreso, educación, desigualdad), no
# porque produzcan el mejor R² a posteriori.
CURATED_INDICATORS = {
    "SP.DYN.LE00.IN": "esperanza_vida",
    "SH.XPD.CHEX.GD.ZS": "gasto_salud_pct_pib",
    "NY.GDP.PCAP.CD": "pib_per_capita_usd",
    "SH.H2O.BASW.ZS": "acceso_agua_pct",
    "SH.STA.BASS.ZS": "acceso_saneamiento_pct",
    "EG.ELC.ACCS.ZS": "acceso_electricidad_pct",
    "SP.DYN.IMRT.IN": "mortalidad_infantil",
    "SI.POV.GINI": "indice_gini",
    "SP.POP.GROW": "crecimiento_poblacional_pct",
    "SL.UEM.TOTL.ZS": "desempleo_pct",
}


def download_wdi_zip() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        print(f"ya existe -> {XLSX_PATH} ({XLSX_PATH.stat().st_size:,} bytes), no se re-descarga")
        return

    print(f"descargando {ZIP_URL} (~80MB, real, Banco Mundial)...")
    resp = requests.get(ZIP_URL, timeout=180)
    resp.raise_for_status()
    ZIP_PATH.write_bytes(resp.content)

    with zipfile.ZipFile(ZIP_PATH) as zf:
        zf.extract("WDIEXCEL.xlsx", RAW_DIR)
    print(f"extraído -> {XLSX_PATH} ({XLSX_PATH.stat().st_size:,} bytes)")


def stream_curated_data_rows() -> list[list]:
    """Recorre la hoja `Data` UNA vez, fila por fila (streaming, `read_only=True`),
    y conserva solo las filas cuyo `Indicator Code` está en `CURATED_INDICATORS`
    -- de 401.394 filas reales a un subconjunto manejable, sin cargar la hoja
    completa en memoria.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Data"]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)  # Country Name, Country Code, Indicator Name, Indicator Code, 1960..2024

    kept = [header]
    for row in rows_iter:
        if row[3] in CURATED_INDICATORS:
            kept.append(row)

    wb.close()
    return kept


def extract_country_dimension() -> list[list]:
    """Lee la hoja `Country` completa (solo 265 filas, liviana) -- la dimensión
    que permite distinguir países reales de agregados regionales/de ingreso
    ("World", "OECD members", "Africa Eastern and Southern"...), que en esta
    hoja tienen `Region` vacío mientras que un país real siempre lo tiene
    poblado.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Country"]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def main() -> None:
    download_wdi_zip()

    curated = stream_curated_data_rows()
    header, data_rows = curated[0], curated[1:]
    import csv
    curated_path = RAW_DIR / "wdi_curated_wide.csv"
    with open(curated_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data_rows)
    print(f"Data (curado): {len(data_rows):,} filas de {10} indicadores -> {curated_path}")

    country_rows = extract_country_dimension()
    country_path = RAW_DIR / "wdi_country_dim.csv"
    with open(country_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(country_rows)
    print(f"Country (dimensión completa): {len(country_rows) - 1:,} filas -> {country_path}")


if __name__ == "__main__":
    main()
